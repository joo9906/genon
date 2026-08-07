"""FAQ → XLSX (openpyxl).

`archive/sfr018-export` 태그의 `export_pipeline/xlsx_export.py` 를 이 배포 단위로
가져온 것이다. 그 브랜치는 FAQ 호출부가 없어 계약만 준비돼 있었다(HANDOFF §3).

FAQ 는 원본 문서를 되쓰는 게 아니라 **표를 새로 만든다.** openpyxl 은 순수 파이썬이라
폐쇄망에서 pip 설치만으로 동작한다 (전처리기에도 openpyxl 이 있지만 그건 xlsx 를
**읽는** 용도다).

3.8절: 로그에 질문·답변 내용을 남기지 않는다. 행 수만 남긴다.
"""

import io

from ..logging_utils import log_info
from .errors import ExportError, ExporterUnavailable, ensure_exportable_items

_HEADERS = ("번호", "질문", "답변", "근거")

# 열 너비 — 답변이 길어 맞춰주지 않으면 읽을 수 없다
_WIDTHS = (6, 40, 80, 40)


def _cell_text(value) -> str:
    """엑셀 셀에 넣을 문자열.

    `=` `+` `-` `@` 로 시작하는 값은 엑셀이 **수식으로 해석한다.** FAQ 답변에 그런
    글자가 올 수 있으므로 앞에 홑따옴표를 붙여 텍스트로 고정한다(수식 인젝션 방지).
    """
    text = str(value if value is not None else "").replace("\r\n", "\n")
    return f"'{text}" if text.startswith(("=", "+", "-", "@")) else text


def build_faq_xlsx(items: list, sheet_title: str = "FAQ") -> bytes:
    """FAQ 항목을 xlsx 바이트로 만든다.

    Args:
        items: `[{"question": str, "answer": str, "sources": str | [str]}, …]`
            (`formatting.to_export_rows` 산출 형태).

    Raises:
        ExporterUnavailable: openpyxl 이 이미지에 없다.
        ExportError: 항목이 없거나 형식이 맞지 않거나 상한을 넘음.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        # 지연 import — 모듈 로드 시점에 import 하면 openpyxl 없는 이미지에서
        # 컨테이너가 아예 뜨지 않는다 (pdf_export 와 같은 이유)
        raise ExporterUnavailable("이 환경에서는 엑셀 내보내기를 사용할 수 없습니다.") from exc

    ensure_exportable_items(items)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = (sheet_title or "FAQ")[:31] or "FAQ"  # 엑셀 시트명 상한 31자

    sheet.append(list(_HEADERS))
    for index, header_cell in enumerate(sheet[1], start=1):
        header_cell.font = Font(bold=True)
        header_cell.alignment = Alignment(vertical="center")
        sheet.column_dimensions[get_column_letter(index)].width = _WIDTHS[index - 1]
    sheet.freeze_panes = "A2"  # 머리행 고정 — 항목이 많아도 열 이름이 보인다

    for position, item in enumerate(items, start=1):
        if not isinstance(item, dict):
            raise ExportError("FAQ 항목 형식이 올바르지 않습니다.")
        question = _cell_text(item.get("question"))
        answer = _cell_text(item.get("answer"))
        if not question.strip() and not answer.strip():
            raise ExportError("질문과 답변이 모두 비어 있는 FAQ 항목이 있습니다.")
        sources = item.get("sources")
        if isinstance(sources, (list, tuple)):
            sources = ", ".join(str(source) for source in sources if str(source).strip())
        sheet.append([position, question, answer, _cell_text(sources)])
        # 줄바꿈이 있는 답변은 wrap 이 없으면 한 줄로 뭉개진다
        for column in (2, 3, 4):
            sheet.cell(row=position + 1, column=column).alignment = Alignment(
                wrap_text=True, vertical="top"
            )

    buffer = io.BytesIO()
    workbook.save(buffer)
    log_info("FAQ 엑셀 생성 완료", event="faq_xlsx_built", item_count=len(items))
    return buffer.getvalue()
